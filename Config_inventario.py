import openpyxl
import os
import random
from prettytable import PrettyTable
from openpyxl.utils import column_index_from_string




def añadir_equipo():
    # Open the Excel file
    wb = openpyxl.load_workbook('DBs/Inventario.xlsx')

    # Select the active worksheet
    ws = wb.active

    # Find the last row of data in the worksheet
    last_row = ws.max_row

    # Find the last column of data in the worksheet
    last_col = ws.max_column

    # Create a list of column headings
    headings = []
    for col in range(1, last_col+1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            headings.append(value)

    # Generate a random and unique ID
    asset_id = random.randint(1000, 9999)
    id_column = column_index_from_string('A')
    used_ids = [ws.cell(row=row, column=id_column).value for row in range(2, last_row+1)]
    while asset_id in used_ids:
        asset_id = random.randint(1000, 9999)

    # Create a new row with the generated ID
    new_row = [str(asset_id)]
    for heading in headings[1:]:
        value = input(f"Enter a value for {heading}: ")
        new_row.append(value)

    # Add the new row to the worksheet
    ws.append(new_row)

    # Save the changes to the Excel file
    wb.save('DBs/Inventario.xlsx')

    print(f"Asset {asset_id} added to inventory")





def eliminar_equipo():
    """Deletes an asset from the inventory"""
    inventory_file = "DBs/Inventario.xlsx"
    try:
        wb = openpyxl.load_workbook(inventory_file)
        ws = wb.active
        inventory = {}
        header_row = ws[1]
        for row in ws.iter_rows(min_row=2, values_only=True):
            asset_id = row[0]
            asset_details = {}
            for i in range(1, len(header_row)):
                header = header_row[i].value
                asset_details[header] = row[i]
            inventory[asset_id] = asset_details
    except FileNotFoundError:
        print("Inventory file not found!")
        return

    asset_id = input("Enter ID to delete: ")
    if asset_id in inventory:
        del inventory[asset_id]  # Remove the item from the dictionary directly
        asset_row = None
        for i, row in enumerate(ws.rows):
            if row[0].value == asset_id:
                asset_row = i + 1
                break
        if asset_row is not None:
            ws.delete_rows(asset_row)
            wb.save(inventory_file)
            print(f"Asset {asset_id} deleted from inventory")
        else:
            print(f"Asset {asset_id} not found in inventory")
    else:
        print(f"Asset {asset_id} not found in inventory")
    print()





def ver_inventario():
    # Open the Excel file
    wb = openpyxl.load_workbook('DBs/Inventario.xlsx')

    # Select the active worksheet
    ws = wb.active
    
    # Get the header row
    header_row = [cell.value for cell in ws[1]]

    # Create a PrettyTable object with the header row
    table = PrettyTable(header_row)

    # Iterate through the rows and add them to the table
    for row in ws.iter_rows(min_row=2, values_only=True):
        table.add_row(row)

    # Print the table
    print(table)

    # Save the changes to the Excel file
    wb.save('DBs/Inventario.xlsx')

    input("\nPress Enter to continue...\n")



def exportar_inventario():
    """Exports the inventory to an Excel file"""
    # Load the Excel file
    try:
        wb = openpyxl.load_workbook('DBs/Inventario.xlsx')
        ws = wb.active
    except FileNotFoundError:
        print("Inventory file not found!")
        return

    # Create a new workbook for the export
    export_wb = openpyxl.Workbook()
    export_ws = export_wb.active
    export_ws.title = "Inventario"

    # Write the header row
    header_row = [cell.value for cell in ws[1]]
    for col, header in enumerate(header_row, start=1):
        export_ws.cell(row=1, column=col, value=header)

    # Write the data rows
    for row, data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        for col, value in enumerate(data, start=1):
            export_ws.cell(row=row, column=col, value=value)

    # Get the export filename
    export_filename = input("Enter the export filename: ") + '.xlsx' # Added the extension
    export_file = os.path.join(os.getcwd(), export_filename)

    # Save the export file
    export_wb.save(export_file)
    print(f"Inventory exported to {export_file}")

