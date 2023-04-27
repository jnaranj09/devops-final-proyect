import openpyxl
import getpass
import os

# Load inventory data from Excel file
inventory_file = "/home/eblanco/Documents/devops-final-proyect/DBs/Inventario.xlsx"
try:
    inventory_workbook = openpyxl.load_workbook(inventory_file)
    inventory_sheet = inventory_workbook.active
    inventory = {}
    header_row = inventory_sheet[1]
    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        asset_id = row[0]
        asset_details = {}
        for i in range(1, len(header_row)):
            header = header_row[i].value
            asset_details[header] = row[i]
        inventory[asset_id] = asset_details
except FileNotFoundError:
    inventory_workbook = openpyxl.Workbook()
    inventory_sheet = inventory_workbook.active
    inventory_sheet.append(["ID"])
    inventory = {}

def ver_inventario():
    """Prints the current inventory"""
    print("Inventory:")
    if not inventory:
        print("No items in inventory")
    for asset_id, asset_details in inventory.items():
        print(f"ID: {asset_id}, {', '.join([f'{k}: {v}' for k,v in asset_details.items()])}")
    print()

def exportar_inventario():
    """Exports the inventory to an Excel file"""
    inventory_workbook = openpyxl.Workbook()
    inventory_worksheet = inventory_workbook.active
    inventory_worksheet.title = "Inventario"

    inventory_worksheet.cell(row=1, column=1, value="ID")
    header_row = inventory_sheet[1]
    for i in range(1, len(header_row)):
        header = header_row[i].value
        inventory_worksheet.cell(row=1, column=i+1, value=header)

    for row, (asset_id, asset_details) in enumerate(inventory.items(), start=2):
        inventory_worksheet.cell(row=row, column=1, value=asset_id)
        for i in range(1, len(header_row)):
            header = header_row[i].value
            inventory_worksheet.cell(row=row, column=i+1, value=asset_details.get(header, ""))

    username = getpass.getuser()
    custom_name = input("Enter a custom name for the file: ")
    filename = f"{custom_name}copy.xlsx"
    inventory_file = os.path.join(os.getcwd(), filename)
    inventory_workbook.save(inventory_file)
    print(f"Inventory exported to {inventory_file}")




def añadir_equipo():
    """Adds a new asset to the inventory"""
    asset_id = input("Enter ID: ")
    inventory[asset_id] = {}  # add this line to create a new entry in the inventory dictionary
    header_row = inventory_sheet[1]
    for i in range(1, len(header_row)):
        header = header_row[i].value
        value = input(f"Enter {header}: ")
        inventory[asset_id][header] = value

    # Append the new asset information to the inventory xlsx file
    row = [asset_id]
    for header in header_row[1:]:
        value = inventory[asset_id][header.value]
        row.append(value)
    inventory_sheet.append(row)
    inventory_workbook.save(inventory_file)

    print(f"Asset {asset_id} added to inventory")


def eliminar_equipo():
    """Deletes an asset from the inventory"""
    asset_id = input("Enter ID to delete: ")
    if asset_id in inventory:
        del inventory[asset_id]
        inventory_workbook.save(inventory_file)
        inventory_sheet.delete_rows(1)
        print(f"Asset {asset_id} deleted from inventory")
    else:
        print(f"Asset {asset_id} not found in inventory")
    print()


eliminar_equipo()



