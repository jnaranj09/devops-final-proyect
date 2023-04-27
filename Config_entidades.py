from variables import ent_key
from openpyxl import load_workbook, utils
from openpyxl.worksheet.table import TableStyleInfo, Table, TableColumn

def crear_entidad():
    # Cargando el archivo
    nombre_entidad = input("Ingrese el nombre de la entidad a crear: ")
    filesheet = "DBs/Inventario.xlsx"
    wb = load_workbook(filesheet)
    sheet = wb.active

    # select the table by name
    table = sheet.tables["Table1"]

    # get the last column of the table
    last_col_index = table.tableColumns[-1].id + 1
    last_col_letter = utils.get_column_letter(last_col_index)
    
    # add the new column header
    sheet[last_col_letter + "1"] = nombre_entidad
    
    # add a new table column to the table
    table.tableColumns.append(TableColumn(id=last_col_index, name=nombre_entidad))
    
    # adjust the table range
    table.ref = f"A1:{chr(sheet.max_column + 64)}{sheet.max_row}"
    
    # style the table
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    
    # save the workbook
    wb.save(filesheet)

    print("\nEntidad creada con exito!\n")




def eliminar_entidad():
    # Cargando el archivo
    filesheet = "DBs/Inventario.xlsx"
    wb = load_workbook(filesheet)
    sheet = wb.active

    # select the table by name and get the range of the table
    table = sheet.tables["Table1"]

    # Mostrando las entidades
    print("Entidades disponibles: ")

    # get a list of the table columns and their corresponding letters
    columns = [(utils.get_column_letter(table.tableColumns.index(column) + 1), column.name) for column in table.tableColumns]

    # print the list of columns and their numbers
    for column_number, column_name in columns:
        print(f"{column_name} = {column_number}.")

    # Seleccionando la entidad a eliminar y eliminandola
    entidad_a_eliminar = input("Ingrese la letra de la entidad que desea eliminar: ")
    ent_index = utils.column_index_from_string(entidad_a_eliminar.upper())
    table.tableColumns.remove(table.tableColumns[ent_index - 1])  #This one removes the table column
    sheet.delete_cols(ent_index) #This one removes the column from the sheet
    table.ref = f"A1:{chr(sheet.max_column + 64)}{sheet.max_row}" #This one adjusts the table range

    # style the table
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    # save the workbook
    wb.save(filesheet)

    print("\nEntidad eliminada con exito!\n")


def ver_entidades():
    # Cargando el archivo
    filesheet = "DBs/Inventario.xlsx"
    wb = load_workbook(filesheet)
    sheet = wb.active

    # select the table by name and get the range of the table
    table = sheet.tables["Table1"]

    # Mostrando las entidades
    print("Entidades disponibles: ")

    # get a list of the table columns and their corresponding letters
    columns = [(utils.get_column_letter(table.tableColumns.index(column) + 1), column.name) for column in table.tableColumns]

    # print the list of columns and their numbers
    for column_number, column_name in columns:
        print(f"{column_name} = {column_number}.")

    # save the workbook
    wb.save(filesheet)

